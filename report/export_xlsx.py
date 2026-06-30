"""XLSX ledger export."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audit.ledger import ledger_rows

HEADERS = [
    "item_id",
    "slide_number",
    "element_type",
    "wcag_criterion",
    "section_508_ref",
    "risk_level",
    "issue_description",
    "suggested_fix",
    "status",
    "applied_at",
    "approved_by",
    "rolled_back_at",
    "snapshot_path",
]


def export_ledger_xlsx(ledger_path: Path, output_path: Path) -> Path:
    """Export the JSON ledger as a readable Excel workbook."""
    rows = ledger_rows(ledger_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Accessibility Ledger"
    ws.append(HEADERS)

    for row in rows:
        ws.append([_cell_value(row.get(header)) for header in HEADERS])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, header in enumerate(HEADERS, start=1):
        max_length = len(header)
        for cell in ws[get_column_letter(column_index)]:
            max_length = max(max_length, len(str(cell.value or "")))
        width = min(max(max_length + 2, 12), 48)
        ws.column_dimensions[get_column_letter(column_index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _cell_value(value: Any) -> str | int | float | None:
    """Convert ledger values into Excel cell-safe scalar values."""
    if value is None:
        return None
    if isinstance(value, (str, int, float)):
        return value
    return str(value)
